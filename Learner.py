import db as dbFile
from word import Word

from copy import deepcopy
import numpy as np
import random
from dataclasses import dataclass, field
from typing import List
import pandas as pd
from TypeNew import Type, TChunk, VChunk
from clusterer import Clusterer

import math

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill

import matplotlib.pyplot as plt

from typing import Tuple, Dict, Any, Optional
from SChunk import SChunk, ChunkPair

import sys
sys.setrecursionlimit(1500)

def softmax(weights: Dict, tau: float = 1.0) -> Dict:
    """Compute softmax distribution with temperature tau."""
    if not weights:
        return {}
    keys, values = zip(*weights.items())
    scaled = [v / tau for v in values]
    max_scaled = max(scaled)  # for numerical stability
    exp_values = [math.exp(s - max_scaled) for s in scaled]
    total = sum(exp_values)
    probs = [e / total for e in exp_values]
    return dict(zip(keys, probs))

def merged_softmax_choice(
    left: Dict, right: Dict, tau: float = 1.0
) -> Tuple[str, Optional[object]]:
    """
    Returns (side, selected_key) where side ∈ {'left', 'right'}.
    Handles empty dicts gracefully.
    """
    if not left and not right:
        return random.choice([('left', None), ('right', None)])

    if not left:
        chosen = softmax_choice(right, tau)
        return 'right', chosen
    if not right:
        chosen = softmax_choice(left, tau)
        return 'left', chosen

    # Both non-empty: proceed with merged softmax
    left_soft = softmax(left, tau)
    right_soft = softmax(right, tau)

    merged_keys = list(left_soft.keys()) + list(right_soft.keys())
    merged_probs = list(left_soft.values()) + list(right_soft.values())

    chosen_key = random.choices(merged_keys, weights=merged_probs, k=1)[0]
    if chosen_key in left_soft:
        return 'left', chosen_key
    else:
        return 'right', chosen_key

def softmax_choice(weights: Dict, tau: float = 1.0):
    """Sample a key from a softmax distribution of weights."""
    probs = softmax(weights, tau)
    keys, values = zip(*probs.items())
    return random.choices(keys, weights=values, k=1)[0]



@dataclass
class LearnerConfig:
    n_trials: int= 10
    border: str = 'next'
    initial_value_chunking: float = -1.
    initial_value_border: float = 1.
    alpha: float = 0.2
    alpha_v: float = 0.3
    beta: float = 1.
    positive_reinforcement: float = 5.
    negative_reinforcement: float = -10.
    RW: bool = False
    chaining: bool = False
    bad_type_threshold: float = -3
    good_type_threshold: float = 2.
    tau: float = 2.0 # softmax parameter of TypeAssigner
    type_on: bool = False
    # parameter for choosing type of learner (RW or not.)

class LongTermMemory():
    
    def __init__(self,config):
        self.initial_value_chunking = config.initial_value_chunking
        self.initial_value_border = config.initial_value_border
        self.behaviour_repertoire = dict() # dictionary of where the keys are couples of chunks and the value a list of behavioural values
        self.chunk_values = dict()
        
        self.chunk_type_associations = dict()
        self.typatory = dict()

    
    def add(self,couple):
        if couple not in self.behaviour_repertoire:
            values = [self.initial_value_border] + [self.initial_value_chunking] * (couple.s1.depth + 1)
            
            self.behaviour_repertoire[couple] =np.array(values)
         
    def update_repertoire(self,couple): # couple must be a couple of SChunks
        subcouples = couple.get_sub_couples()
        for c in subcouples:
            self.add(c)
            
    def update_chunk(self,chunk):
        if chunk not in self.chunk_values:
            self.chunk_values[chunk] = 0.0
            
    def update_typatory(self,ttype):
        if ttype not in self.typatory:
            self.typatory[ttype] = 0.0
            
    def update_chunk_type_associations(self,chunk,ttype):
        if chunk not in self.chunk_type_associations:
            self.chunk_type_associations[chunk] = dict()
        if ttype not in self.chunk_type_associations[chunk]:
            self.chunk_type_associations[chunk][ttype] = 0.0
            
    def decay_chunk_type_values(self):
        multiplier = 0.999
        for c in self.chunk_type_associations:
            for t in self.chunk_type_associations[c]:
                self.chunk_type_associations[c][t] *= multiplier
                
                
    def clean_chunk_type_associations(self):
        elements_to_clean = []
        for c in self.chunk_type_associations:
            for t in self.chunk_type_associations[c]:
                if np.abs(self.chunk_type_associations[c][t]) < 0.5:
                    elements_to_clean.append((c,t))
                    
        for (c,t) in elements_to_clean:
            self.chunk_type_associations[c].pop(t)
        
    def display_typings_of_elements(self):
        for c,d in self.chunk_type_associations.items():
            if len(c) ==1:
                print(f'The type of {c} are {d}')
            



    def write_behaviour_repertoire_to_xlsx(self, filename="structured_output.xlsx"):
        # Extract key attributes and values
        rows = []
        for key, values in self.behaviour_repertoire.items():
            for i, value in enumerate(values):
                rows.append({
                    's1': repr(key.s1),
                    's2': repr(key.s2),
                    'length': len(key.s1),
                    'index': i,
                    'value': value
                })
    
        # Create DataFrame
        df = pd.DataFrame(rows)
    
        # Pivot to wide format
        df_pivoted = df.pivot_table(
            index=['length', 's1', 's2'],
            columns='index',
            values='value',
            aggfunc='first'
        ).reset_index()
    
        # Sort by length, s1, s2
        df_pivoted = df_pivoted.sort_values(by=['length', 's1', 's2'])
    
        # Write to Excel
        df_pivoted.to_excel(filename, index=False)
    
        # Post-process Excel file with openpyxl
        wb = load_workbook(filename)
        ws = wb.active
    
        # Bold headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
    
        # Freeze top row
        ws.freeze_panes = 'A2'
        
    
        # Define fill color for max value
        highlight_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold
        
        # Find the first data row (assumed row 2)
        first_data_row = 2
        last_row = ws.max_row
        first_data_col = 4  # Adjust if your "value" columns start later
        last_col = ws.max_column
        
        # Iterate through rows and highlight the max value
        for row in ws.iter_rows(min_row=first_data_row, max_row=last_row,
                                min_col=first_data_col, max_col=last_col):
            # Extract values and find max
            values = [cell.value for cell in row if isinstance(cell.value, (int, float))]
            if not values:
                continue
            max_val = max(values)
        
            # Highlight all cells with the max value
            for cell in row:
                if cell.value == max_val:
                    cell.fill = highlight_fill

    
        wb.save(filename)

            
@dataclass
class LearningHistory:
    success: List[int] = field(default_factory=list)
    sent_len: List[int] = field(default_factory=list)

    def record(self, success_value: int, length: int):
        self.success.append(success_value)
        self.sent_len.append(length)
        


    def plot_moving_average(self, window_size=10, show=True, save_path=None):
        if len(self.success) < window_size:
            print(f"Not enough data to compute moving average (need at least {window_size}).")
            return

        ma = [sum(self.success[i:i+window_size]) / window_size 
              for i in range(len(self.success) - window_size + 1)]

        plt.figure(figsize=(10, 5))
        plt.plot(ma, label=f'{window_size}-trial moving avg')
        plt.xlabel('Trial')
        plt.ylabel('Success rate')
        plt.ylim((0,1))
        plt.title('Learning Progress')
        plt.grid(True)
        plt.legend()

        if save_path:
            plt.savefig(save_path)
        if show:
            plt.show()
        plt.close()
        return ma
        

    
    def plot_moving_average_by_length_timed(self, window_size=10, show=True, save_path=None, lengths=None):
        data = pd.DataFrame({
            'success': self.success,
            'length': self.sent_len
        })
    
        if lengths is None:
            lengths = sorted(data['length'].unique())
    
        plt.figure(figsize=(10, 5))
        learning_curves = {}
    
        for length in lengths:
            # Create a time-aligned series: NaN for trials not matching the length
            mask = data['length'] == length
            successes = data['success'].where(mask)
    
            # Compute moving average ignoring NaNs
            ma = successes.rolling(window=window_size, min_periods=1).mean()
    
            plt.plot(ma, label=f'len={length}')
            
            # Store the aligned success list (not the moving avg) for averaging across learners
            learning_curves[length] = list(successes)
    
        plt.xlabel('Trial (global time)')
        plt.ylabel('Success rate')
        plt.title(f'Learning Progress by Sentence Length (Aligned in Time, window={window_size})')
        plt.grid(True)
        plt.legend()
    
        if save_path:
            plt.savefig(save_path)
        if show:
            plt.show()
        plt.close()
        
        return learning_curves

            
class WorkingMemory():
    def __init__(self,learner,config: LearnerConfig,samplenum,cfg,output=False,clustering=False):
        self.learner = learner
        self.reinforcer = Reinforcer(learner,config)
        self.type_assigner = TypeAssigner(learner,config)
        self.events = []
        self.typing_events = dict()
        self.ts1 = TChunk(Type.EMPTY)
        self.ts2 = TChunk(Type.EMPTY)
        self.typing_used = False
        self.border_before = True
        self.border_within = False
        self.border_type = config.border
        self.beta = config.beta
        self.pos = config.positive_reinforcement
        self.neg = config.negative_reinforcement

        self.success = []
        self.output = output

        # print(type(learner), type(config), type(samplenum), type(cfg))
        if samplenum == 0:
            self.db = dbFile.DB(clustering=clustering)
        else:
            self.db = dbFile.DB(clustering=clustering, clusters=Clusterer(cfg).cluster(samplenum))
        self.deps_down = dict()
        self.deps_up   = dict()
        self.lens = []

    def get_responses(self):
        responses = []
        for e in self.events:
            responses.append(e[1])
        return responses
    
    def respond(self,stimuli_stream,s1,s2_index,reinforcement = True):
        # get the s2 stimuli and make it a chunk
        try:
            word = Word(stimuli_stream.read_stimuli(s2_index), self.db)
            s2 = SChunk(word.t)
        except IndexError:
            sys.exit("Index doesn't exist. End of input reached before learning is finished.")
        
        # print(len(self.ws), str(s1).count(',')+1)
        pair = ChunkPair((s1,s2))

        # Update memory
        self.learner.ltm.update_repertoire(pair)

        response = self.choose_behaviour(pair)

        self.events.append((pair,response))
        # print(self.get_responses())
        
        if response == 0: # boundary placement
            self.learner.n_reinf += 1
            self.lens.append(len(self.ws))

            self.custom_reinforce(self.is_border_correct(stimuli_stream, s2_index))
            self.ws = [word]
            self.deps_down = dict()
            self.deps_up   = dict()

            sent_length = stimuli_stream.length_current_sent(s2_index - 1)
            if self.is_border_correct(stimuli_stream,s2_index):
                self.success.append(1)
                reward = self.pos
                self.learner.history.record(1,sent_length)
                if reinforcement:
                    self.reinforcer.reinforce2(self.events,reward)
                    #self.reinforcer.reinforce_value_hierarchical(pair.s1,reward)
            else:
                reward = self.neg
                self.success.append(0)
                self.learner.history.record(0,sent_length)
                if reinforcement:
                    self.reinforcer.reinforce2(self.events,reward)
                    #self.reinforcer.reinforce_value(pair.s1,reward)
            
            new_s1, s2_index = self.get_new_s1(stimuli_stream,s2_index, s2)
            
            self.events = []
               
        else: # some type of chunking occurs
            self.add_dependencies(word)
            new_s1, s2_index = self.chunk(pair, response, stimuli_stream,s2_index) 
  
        return new_s1, s2_index
    
    def respond_with_type(self,stimuli_stream,s1,s2_index,reinforcement = True):
        # get the s2 stimuli and make it a chunk
        try:
            s2 = SChunk(stimuli_stream.read_stimuli(s2_index))
        except IndexError:
            sys.exit("Index doesn't exist. End of input reached before learning is finished.")
        
        pair = ChunkPair((s1,s2))
        self.learner.ltm.update_repertoire(pair)
        
        self.type_assigner.assign_type(pair) # self.ts1 is a TChunk, while ts2 is a TChunk 

        
        # Check if the structure of ts1 and ts2 are useful as a support for decisions
        # If so, use them in the decision making process
        # Otherwise, fallback on the decision making process without types
        if not self.ts1.is_consistent():  
            self.typing_used = False
            response = self.choose_behaviour(pair)
        else: 
            self.typing_used = True
            # I need to fix the following function! Needs to gather support for decisions...
            response = self.choose_behaviour_with_types(pair) # Set also whether self.typing_used is True or False

        self.events.append((pair,response))
        #print(self.get_responses())
        
        if False:
            print(self.ts1)
            print(self.ts1.is_consistent())
            print(self.ts2)
            if response == 0:
                print('border')
            else:
                print('chunk')
        
        
        
        
        if response == 0: # boundary placement
            self.learner.n_reinf += 1
            self.learner.ltm.decay_chunk_type_values()
            if self.learner.n_reinf % 100 == 0:
                self.learner.ltm.clean_chunk_type_associations()
            sent_length = stimuli_stream.length_current_sent(s2_index - 1)
            
            if self.is_border_correct(stimuli_stream,s2_index):
                self.success.append(1)
                reward = self.pos
                self.learner.history.record(1,sent_length)
                if reinforcement:
                    self.reinforcer.reinforce2(self.events,reward)
                    if not self.typing_used:
                       self.type_assigner.type_sentence(pair.s1)
                       self.reinforcer.reinforce_types(self.typing_events,reward)
                    elif self.ts1.reduce() != Type.SENTENCE:
                        self.typing_events = self.extract_typing_events(s1,self.ts1)
                        self.reinforcer.reinforce_types(self.typing_events,self.neg)
                    else:
                        self.typing_events = self.extract_typing_events(s1,self.ts1)
                        self.reinforcer.reinforce_types(self.typing_events,reward)
            else:
                reward = self.neg
                self.success.append(0)

                self.learner.history.record(0,sent_length)
                if reinforcement:
                    self.reinforcer.reinforce2(self.events,reward)
                    if self.typing_used and self.ts1.reduce() == Type.SENTENCE:
                       self.typing_events = self.extract_typing_events(s1,self.ts1)
                       self.reinforcer.reinforce_types(self.typing_events,reward)
            
            new_s1, s2_index = self.get_new_s1(stimuli_stream,s2_index, s2)
            
            self.events = []
            self.typing_events = {}
            self.typing_used = False
               
        else: # some type of chunking occurs
            new_s1, s2_index = self.chunk(pair, response, stimuli_stream,s2_index) 
            
        
        return new_s1, s2_index
    
    def is_border_correct(self,stimuli_stream,s2_index):
        is_border = stimuli_stream.border_before[s2_index]
        return is_border and not self.border_within and self.border_before
    
    def chunk(self,pair,response,stimuli_stream,s2_index):
        if not self.border_within:
            self.border_within = stimuli_stream.border_before[s2_index]
         
        # Perform chunking at correct level
        new_s1 = pair.s1.chunk_at_depth(pair.s2,depth=pair.s1.depth+1-response)
        self.ts1 = self.ts1.chunk_at_depth(self.ts2,depth=pair.s1.depth+1-response)
        s2_index+=1 
        return new_s1, s2_index
    
    def get_new_s1(self,stimuli_stream,s2_index,s2):
        if self.border_type == 'next':
            new_s1,s2_index = stimuli_stream.next_beginning_sent(s2_index)
            new_s1 = SChunk(new_s1)
            self.ts1 = TChunk(Type.EMPTY)
            self.ts2 = TChunk(Type.EMPTY)
        else:
            self.border_before = stimuli_stream.border_before[s2_index]
            new_s1,s2_index = s2, s2_index + 1
            self.ts1 = self.ts2
            self.ts2 = TChunk(Type.EMPTY)

        self.border_within = False
        return new_s1, s2_index
        
    def extract_typing_events(self,s1,ts1,mapping=None):
        if mapping is None:
            mapping = {}
            
        schunk = s1
        tchunk = ts1
    
        # Base case: if it's a leaf node (not a list), just map it
        if not isinstance(schunk.structure, list):
            mapping[schunk] = tchunk.structure
            return mapping
    
        # Map the current (composite) chunk
        mapping[schunk] = tchunk.reduce()
    
        # Recurse on left and right subchunks
        s_left = schunk.get_left()
        s_right = schunk.get_right()
        t_left = tchunk.get_left()
        t_right = tchunk.get_right()
    
        self.extract_typing_events(s_left, t_left, mapping)
        self.extract_typing_events(s_right, t_right, mapping)
    
        return mapping
        # Use the structure of self.ts1 to extract the typing events
        
    def respond_with_chaining(self,stimuli_stream,s1,s2_index,reinforcement = True):
        # Positive and negative propagation to chunks
        # get the s2 stimuli and make it a chunk
        try:
            s2 = SChunk(stimuli_stream.read_stimuli(s2_index))
        except IndexError:
            sys.exit("Index doesn't exist. End of input reached before learning is finished.")
        
        pair = ChunkPair((s1,s2))
        self.learner.ltm.update_repertoire(pair)

        response = self.choose_behaviour(pair)

        event = [(pair,response)]
         
        if response == 0: # boundary placement
            self.learner.n_reinf += 1
            sent_length = stimuli_stream.length_current_sent(s2_index - 1)
            if self.is_border_correct(stimuli_stream,s2_index):
                reward = self.pos
                self.learner.history.record(1,sent_length)
            else:
                reward = self.neg
                self.learner.history.record(0,sent_length)
            
            new_s1, s2_index = self.get_new_s1(stimuli_stream,s2_index, s2)
               
        else: # some type of chunking occurs
            # Check if there was a border
            new_s1, s2_index = self.chunk(pair, response, stimuli_stream,s2_index) 
            self.learner.ltm.update_chunk(new_s1)
            reward = self.learner.ltm.chunk_values[new_s1]
            
        # Reinforce the event and the value of s1.
        if reinforcement:
            self.reinforcer.reinforce2(event,reward)
            self.reinforcer.reinforce_value(pair.s1,reward)
            
        return new_s1, s2_index
    
    def respond_with_chaining2(self,stimuli_stream,s1,s2_index,reinforcement = True):
        # Only positive propagation to chunks
        
        # get the s2 stimuli and make it a chunk
        try:
            s2 = SChunk(stimuli_stream.read_stimuli(s2_index))
        except IndexError:
            sys.exit("Index doesn't exist. End of input reached before learning is finished.")
        
        pair = ChunkPair((s1,s2))
        self.learner.ltm.update_repertoire(pair)

        response = self.choose_behaviour(pair)

        event = [(pair,response)]
        
        if response == 0: # boundary placement
            self.learner.n_reinf += 1
            sent_length = stimuli_stream.length_current_sent(s2_index - 1)

            if self.is_border_correct(stimuli_stream,s2_index):
                reward = self.pos
                self.learner.history.record(1,sent_length)
                if reinforcement:
                    self.reinforcer.reinforce2(event,reward)
                    self.reinforcer.reinforce_value_hierarchical(pair.s1,reward)
            else:
                reward = self.neg
                self.learner.history.record(0,sent_length)
                if reinforcement:
                    self.reinforcer.reinforce2(event,reward)
                    #self.reinforcer.reinforce_value(pair.s1,reward)
            
            new_s1, s2_index = self.get_new_s1(stimuli_stream,s2_index, s2)
            
        else: # some type of chunking occurs
            # Check if there was a border
            new_s1, s2_index = self.chunk(pair, response, stimuli_stream,s2_index) 
            self.learner.ltm.update_chunk(new_s1)
            reward = self.learner.ltm.chunk_values[new_s1]
            if reinforcement:
                self.reinforcer.reinforce2(event,reward)
                self.reinforcer.reinforce_value_hierarchical(pair.s1,reward)
            
        return new_s1, s2_index
    
    def do_output(self):
        return self.output and self.learner.n_trials-self.learner.n_reinf<200

    def custom_reinforce(self, positive):
        # print("prefill | ", end='')
        # print(self.ws, self.deps_down)
        self.fill_deps()

        if positive:
            if self.do_output():
                # print("postfill | ", end='')
                print(self.ws, self.deps_down)
                # print("----------------------------------------")
            u = 1
        else:
            u = -1

        for i in range(len(self.ws)):
            numDeps = dict()
            # left
            for j in range(i-1,-1,-1):
                if self.ws[i].t==self.ws[j].t:
                    continue

                key = (self.ws[i].t,self.ws[j].t,Dir.L)
                if key not in numDeps: numDeps[key]=0
                if i in self.deps_down and j in self.deps_down[i]:
                    numDeps[key] += 1

            # right
            for j in range(i+1,len(self.ws)):
                if self.ws[i].t==self.ws[j].t:
                    continue

                key = (self.ws[i].t,self.ws[j].t,Dir.R)
                if key not in numDeps: numDeps[key]=0
                if i in self.deps_down and j in self.deps_down[i]:
                    numDeps[key] += 1

            for k in numDeps:
                self.db.reinforce(self.ws[i].name, k, numDeps[k], u)

    def fill_deps(self):
        # print(f"Filling dependencies for sentence {self.ws} | {self.deps_down}")

        # select random root
        heads = list(filter(lambda x:self.is_head_free(x) and (x not in self.deps_up or self.deps_up[x]==None), range(len(self.ws))))
        if len(heads) == 0:
            # print("Only one head, sentence is all satisfied")
            return
        # print(f"heads: {heads}")
        root = random.choice(heads)

        for i in range(len(self.ws)):
            # ignore every word which isn't a head
            if (i in self.deps_up and self.deps_up[i]!=None) or i==root:
                continue
            # print(f"Word i: {i} | root {root}")

            cands = set()

            # iter to left
            j = i-1
            # print(f"Decrementing j from {j}")
            while j>=0:
                # print(f"j: {j} | ancestor: {self.get_ancestor(j)} | num_deps: {self.ws[j].get_num_dep_type(self.ws[i].t, Dir.R, self.db)}")
                if self.get_ancestor(j)!=i and self.ws[j].t != self.ws[i].t and self.ws[j].get_num_dep_type(self.ws[i].t, Dir.R, self.db)<2:
                    # print("adding")
                    cands.add(j)

                # if a dep goes over i from j then anything left of j makes xdep
                if root == j:
                    break
                elif self.goes_over_to_right(j,i):
                    break
                else:
                    j = self.get_next_left_cand(i,j)

            # !!! can't link to more than 2 of the same type

            # iter to right
            j = i+1
            # print(f"Incrementing j from {j}")
            while j<len(self.ws):
                # print(f"j: {j} | ancestor: {self.get_ancestor(j)} | num_deps: {self.ws[j].get_num_dep_type(self.ws[i].t, Dir.L, self.db)}")
                if self.get_ancestor(j)!=i and self.ws[j].t != self.ws[i].t and self.ws[j].get_num_dep_type(self.ws[i].t, Dir.L, self.db)<2:
                    # print("adding")
                    cands.add(j)

                if root == j:
                    break
                elif self.goes_over_to_left(j,i):
                    break
                else:
                    j = self.get_next_right_cand(i,j)
            
            # select best cand
            candidates = list(map(lambda x:(x,self.dep_utility_increase(x,i)), cands))
            if len(candidates) > 0:
                iChosen = max(candidates, key=lambda x:x[1])[0]
                # print(f"Chosen {iChosen}->{i}")
                self.add_dep(iChosen, i)

        # self.ws[iChosen].increment_dependency_count(self.ws[i].t, Dir.L if i < iChosen else Dir.R, self.db)

        # print(f"Filled deps | {self.ws} | {self.deps_down}")

    def get_ancestor(self, i):
        if i not in self.deps_up or self.deps_up[i] == None:
            return i
        else:
            return self.get_ancestor(self.deps_up[i])

    def is_head_free(self, i):
        for j in range(len(self.ws)):
            if j==i: continue
            
            if j in self.deps_up and self.deps_up[j]!=None:
                if min(self.deps_up[j], j) < i and i < max(self.deps_up[j],j):
                    return False
            
        return True

    def goes_over_to_right(self, j, i):
        if j in self.deps_up and self.deps_up[j]!=None and i<self.deps_up[j]:
            return True
        
        if j in self.deps_down and len(self.deps_down[j])>0 and max(self.deps_down[j])>i:
            return True

        return False

    def get_next_left_cand(self, i, j):
        if j in self.deps_up and self.deps_up[j]!=None and j>self.deps_up[j]:
            return self.deps_up[j]
        
        if j in self.deps_down and len(self.deps_down[j])>0:
            return min(j-1, min(self.deps_down[j])-1)

        return j-1

    def goes_over_to_left(self, j, i):
        if j in self.deps_up and self.deps_up[j]!=None and i>self.deps_up[j]:
            return True
        
        if j in self.deps_down and len(self.deps_down[j])>0 and min(self.deps_down[j])<i:
            return True

        return False

    def get_next_right_cand(self, i, j):
        if j in self.deps_up and self.deps_up[j]!=None and j<self.deps_up[j]:
            return self.deps_up[j]
        
        if j in self.deps_down and len(self.deps_down[j])>0:
            return max(j+1, max(self.deps_down[j])+1)

        return j+1
    
    def get_ldeps(self, i):
        if i not in self.deps_down or len(self.deps_down[i])==0:
            return set([i])
        
        lmostdep = min(self.deps_down[i])
        if lmostdep > i:
            return set([i])
        
        else:
            lchildren = self.get_ldeps(lmostdep)
            lchildren.add(i)
            return lchildren
    
    def get_rdeps(self, i):
        if i not in self.deps_down or len(self.deps_down[i])==0:
            return set([i])
        
        rmostdep = max(self.deps_down[i])
        if rmostdep < i:
            return set([i])
        
        else:
            rchildren = self.get_rdeps(rmostdep)
            rchildren.add(i)
            return rchildren

    def add_dependencies(self, w):
        self.ws.append(w)
        # print(f"\n$$$$$$$$$$ Adding deps to sentence | {self.ws} | {self.deps_down} $$$$$$$$$$")
        banned_pos_list = self.compute_banned_pos(w)
        # print(banned_pos_list)

        # make deps from new word to left elems
        for i in range(len(self.ws)-2,-1,-1):
            if i in banned_pos_list or self.get_head(i)!=None:
                continue
            elif self.does_dep_increase_utility(len(self.ws)-1, i) and not self.is_dep_impossible(i, len(self.ws)-1):
                # print(f"Adding dep {len(self.ws)-1}->{i} | pre: {self.ws[-1].get_num_dep_type(self.ws[i].t, Dir.L, self.db)}")
                self.add_dep(len(self.ws)-1,i)
                self.ban_descendants(banned_pos_list, i)

        # make deps from an old word to new word
        candidates = set()
        for i in range(len(self.ws)-2,-1,-1):
            if i in banned_pos_list:
                continue
            if self.does_dep_increase_utility(i, len(self.ws)-1) and not self.is_dep_impossible(i, len(self.ws)-1):
                # print(f"Valid candidate")
                candidates.add((i,self.dep_utility_increase(i,len(self.ws)-1)))
        if len(candidates) > 0:
            v = max(candidates, key=lambda x:x[1])[0]
            # print(f"Adding dep {v}->{len(self.ws)-1}")
            self.add_dep(v, len(self.ws)-1)

        # print(f"Added | {self.ws} | {self.deps_down}")

    def is_dep_impossible(self, iH, iD):
        return (iH in self.deps_up and self.deps_up[iH] != None and ((
                self.deps_up[iH] > iH
                and self.deps_up[iH] < iD
            ) or (
                self.deps_up[iH] < iH
                and self.deps_up[iH] > iD
            )))

    def does_dep_increase_utility(self, iH, iD):
        # print(f"{iH}->{iD} | ", end='')
        improvement = self.ws[iH].get_improvement_value(self.ws[iD].t, Dir.L if iD<iH else Dir.R, self.db)
        # print(f"improvement: {improvement}")
        return improvement > 0
        
    def dep_utility_increase(self, iH, iD):
        improvement = self.ws[iH].get_improvement_value(self.ws[iD].t, Dir.L if iD<iH else Dir.R, self.db)
        # print()
        # print(f"improvement: {improvement}")
        return improvement
        
    def compute_banned_pos(self, w):
        # compute cross deps and words of the same type
        banned = set()
        i = 0
        while i<len(self.ws)-1:
            # ban if they have the same type
            if w.t == self.ws[i].t:
                banned.add(i)
            
            # all the ones inbetw i and j must be xdeps
            j = self.get_next_word_pos(i)
            banned = banned.union(range(i+1,j))

            i = max(j,i+1)
        return banned

    def get_next_word_pos(self, i):
        children = self.get_children(i)
        head     = self.get_head(i)
        if len(children) == 0 and head == None:
            return i+1
        elif len(children) == 0:
            return max(i+1,head)
        elif head == None:
            return max(children)
        else:
            return max(max(children), head)

    def ban_descendants(self, banned, i):
        banned.add(i)
        for child in self.get_children(i):
            self.ban_descendants(banned, child)
        
    def get_children(self, i):
        if i not in self.deps_down:
            self.deps_down[i] = []
        return self.deps_down[i]

    def get_head(self, i):
        if i not in self.deps_up:
            self.deps_up[i] = None
        return self.deps_up[i]

    def add_dep(self, iH, iD):
        self.deps_up[iD] = iH
        if iH not in self.deps_down: self.deps_down[iH] = []
        self.deps_down[iH].append(iD)
        self.ws[iH].increment_dependency_count(self.ws[iD].t, Dir.L if iD < iH else Dir.R, self.db)

    def choose_behaviour(self,couple):
        b_range = len(self.learner.ltm.behaviour_repertoire[couple])
        z = self.Q_tilde(couple,b_range)
        weights = np.exp(self.beta * z)
        options = [i for i in range(b_range)]
        response = random.choices(options,weights/np.sum(weights))
        return response[0]  
    
    def get_right_values(self,couple):
        list_right_types = self.ts1.right_types()
        t2= self.ts2.structure 
        
        list_types = self.ts1.remove_structure2()
        list_chunk = couple.s1.flatten_structure()
        list_values = []
        for c,t in zip(list_chunk,list_types):
            self.learner.ltm.update_chunk_type_associations(SChunk(c), t)
            list_values.append(self.learner.ltm.chunk_type_associations[SChunk(c)][t])

        responses = self.get_responses()
            
        value_chunk = VChunk.from_list_and_responses(list_values, responses)
        right_values = value_chunk.right_values()
        return right_values
    
    def get_z_values_type(self,couple):
        list_right_types = self.ts1.right_types()
        t2= self.ts2.structure 
        
        list_types = self.ts1.remove_structure2()
        list_chunk = couple.s1.flatten_structure()
        list_values = []
        for c,t in zip(list_chunk,list_types):
            self.learner.ltm.update_chunk_type_associations(SChunk(c), t)
            list_values.append(self.learner.ltm.chunk_type_associations[SChunk(c)][t])

        responses = self.get_responses()
            
        value_chunk = VChunk.from_list_and_responses(list_values, responses)
        right_values = value_chunk.right_values()
        #print('-----------')
        #print(right_values)
        # Here, I need to get the elements of s1 and s2
        # Construct the VChunk associated, using the chunk_type_associations dictionary
        # Check whether the action is supported and if so, use the correct value to update z
        
        z = np.zeros(len(list_right_types)+1)
        #print(len(z))
        
        reduced_type = list_right_types[-1]
        if reduced_type == Type.SENTENCE and not reduced_type.is_compatible(t2):
            # print('Support for border, s1 well-typed')
            z[0] = right_values[-1]
        
        for i in range(len(list_right_types)):
            t1 = list_right_types[i]
            if t1.is_compatible(t2):
                pass
                # print('Support for chunking')
                z[i+1]=right_values[i]
        return z
    
    def get_value_chunk(self,couple):
        list_types = self.ts1.remove_structure2()
        list_chunk = couple.s1.flatten_structure()
        list_values = []
        for c,t in zip(list_chunk,list_types):
            self.learner.ltm.update_chunk_type_associations(SChunk(c), t)
            list_values.append(self.learner.ltm.chunk_type_associations[SChunk(c)][t])

        responses = self.get_responses()
            
        value_chunk = VChunk.from_list_and_responses(list_values, responses)
        return value_chunk


    def Q_tilde(self,couple,b_range):
        z = deepcopy(self.learner.ltm.behaviour_repertoire[couple])
        subpairs = couple.get_sub_couples()
        
        norm_vec = np.array([b_range - 1]+[i for i in range(b_range-1,0,-1)])
        # Accumulate support from subchunks
        for pair in subpairs:
            lenp = len(self.learner.ltm.behaviour_repertoire[pair])
            z[:lenp] += self.learner.ltm.behaviour_repertoire[pair]
        # Take the average
        z /= norm_vec
        return z 

    def choose_behaviour_with_types(self, couple):
        
        b_range = len(self.learner.ltm.behaviour_repertoire[couple])
        z = self.Q_tilde(couple,b_range)
        z_type = self.get_z_values_type(couple)
        # combine z and z_type with some rules
        z = (z + z_type)/2
        weights = np.exp(self.beta * z)
        options = [i for i in range(b_range)]
        response = random.choices(options,weights/np.sum(weights))
        return response[0]
            
class Reinforcer():
    
    def __init__(self,learner,config: LearnerConfig):
        self.alpha = config.alpha
        self.alpha_v = config.alpha_v
        self.positive_reinforcement = config.positive_reinforcement
        self.negative_reinforcement = config.negative_reinforcement
        self.learner = learner
        self.RW = config.RW
        
    def __repr__(self):
        return f"Reinforcer of {self.learner}"
    
    def get_sub_events(self,event):
        couple, r = event
        subevents = []

        #subevents.append((couple,r))
        for subcouple in couple.get_sub_couples():
            if r < len(self.learner.ltm.behaviour_repertoire[subcouple]):
                subevents.append((subcouple,r))        
        return subevents
            
    def get_sub_eventsRW(self,event):
        couple, r = event
        subevents = []
        
        Q = self.learner.ltm.behaviour_repertoire[couple][r]
        #subevents.append((couple,r))
        for subcouple in couple.get_sub_couples():
            if r < len(self.learner.ltm.behaviour_repertoire[subcouple]):
                subevents.append((subcouple,r)) 
                Q += self.learner.ltm.behaviour_repertoire[subcouple][r]
        return subevents, Q
        
        
    def reinforce(self,events, reinforcement = 'positive'):
        if reinforcement == 'positive':
            u = self.positive_reinforcement
        elif reinforcement == 'negative':
            u = self.negative_reinforcement

        for couple,r in events:
            if self.RW:
                subevents, Q = self.get_sub_eventsRW((couple,r))
                
                for p,rr in subevents:
                    self.learner.ltm.behaviour_repertoire[p][rr]+= self.alpha * (u - Q)
            else:
                subevents=self.get_sub_events((couple,r))
    
                for p,rr in subevents:
                    self.learner.ltm.behaviour_repertoire[p][rr] += self.alpha * (u - self.learner.ltm.behaviour_repertoire[p][rr])
                    
    def reinforce2(self,events, reward):
        u = reward

        for couple,r in events:
            if self.RW:
                subevents, Q = self.get_sub_eventsRW((couple,r))
                
                for p,rr in subevents:
                    self.learner.ltm.behaviour_repertoire[p][rr]+= self.alpha * (u - Q)
            else:
                subevents=self.get_sub_events((couple,r))
    
                for p,rr in subevents:
                    self.learner.ltm.behaviour_repertoire[p][rr] += self.alpha * (u - self.learner.ltm.behaviour_repertoire[p][rr])

    def reinforce_value(self,chunk,reward):
        self.learner.ltm.update_chunk(chunk)
        self.learner.ltm.chunk_values[chunk] += self.alpha_v * (reward - self.learner.ltm.chunk_values[chunk])
        
    def reinforce_value_hierarchical(self,chunk,reward):
        chunks_list = [chunk]
        try:
            subchunks = chunk.get_right_subchunks2(0)
            for s in subchunks:
                chunks_list.append(s)
        except ValueError:
            pass 
        
        for c in chunks_list:
            self.learner.ltm.update_chunk(c)
            self.learner.ltm.chunk_values[c] += self.alpha_v * (reward - self.learner.ltm.chunk_values[c])
            
    def reinforce_types(self,typing_events,reward):
        for chunk,typ in typing_events.items():
            self.learner.ltm.update_chunk_type_associations(chunk,typ)
            self.learner.ltm.chunk_type_associations[chunk][typ] += self.alpha * (reward - self.learner.ltm.chunk_type_associations[chunk][typ])

class TypeAssigner(): #här ska jag vara för att fixa
    
    def __init__(self, learner, config: LearnerConfig):
        self.learner = learner
        self.bad_type_threshold = config.bad_type_threshold
        self.good_type_threshold = config.good_type_threshold
        self.tau = config.tau
        self.type_dict = dict()
        pass

    def add_to_type_dict(self, tok, typ):
        tok=str(tok)
        typ=str(typ)
        # print(f"assigning type {typ} to {tok}")
        if tok in self.type_dict:
            if typ in self.type_dict[tok]:
                self.type_dict[tok][typ]+=1
            else:
                self.type_dict[tok][typ]=1
        else:
            self.type_dict[tok] = dict()
            self.type_dict[tok][typ]=1

    def replace_in_type_dict(self, tok, oldtyp, typ):
        # print(f"reassignment {tok}, {oldtyp}, {typ}")
        return
        tok=str(tok)
        oldtyp=str(typ)
        typ=str(typ)
        print(f"removing type {oldtyp} from {tok}. tok: {self.type_dict[tok]}")
        self.type_dict[tok][oldtyp]-=1
        self.add_to_type_dict(tok, typ)

    
    def assign_type(self, pair: ChunkPair):
        # Do type assignment taking into account the values associated to chunk and types
        left_candidates = self.extract_good_starting_types(pair.s1)
        right_candidates = self.extract_good_types(pair.s2)
        # if pair.s2 in self.learner.ltm.chunk_type_associations:
        #     print("chunk-type-associations for s2 in assign types",self.learner.ltm.chunk_type_associations[pair.s2])
        # else: print("s2 not in chunk-type associations")
        # print("right candidates",right_candidates)
        
        
        if self.learner.wm.ts1.has_empty_elements():
            if isinstance(self.learner.wm.ts1.structure,list): #if the structure of ts1 is a list, it is a compound, thus s1 is inherited
                #print('bad TCHUNK... Assign ts2 to its best candidate (in case it is used as the beginning of the next sentence)')
                if right_candidates:
                    choice = softmax_choice(right_candidates,tau = self.tau)
                    self.learner.wm.ts2 = TChunk(choice)
                    self.add_to_type_dict(pair.s2,choice)
            else:
                #print('Here I should try to assign t1 and t2 jointly')
                if right_candidates:
                    # Here I need to check for consistency
                    choice = softmax_choice(right_candidates,tau = self.tau)
                    self.learner.wm.ts2 = TChunk(choice)
                    self.add_to_type_dict(pair.s2,choice)
                if left_candidates:
                    # Here I need to check for consistency
                    choice = softmax_choice(left_candidates,tau = self.tau)
                    self.learner.wm.ts1 = TChunk(choice)
                    self.add_to_type_dict(pair.s1,choice)

        else:
            #print('s1 typed')
            # if isinstance(self.learner.wm.ts1.structure,list): #if s1 is inherited and typed
            #     #this part made by Anna gets the recursively averaged value of the reduced type for compound s1 and puts them in right candidates
            #     vchunk_s1=self.learner.wm.get_value_chunk(pair.s1) #I added .s1 after talking to jerome, it makes no sense to me to work with the entire pair?
            #     value_ts1 = vchunk_s1.reduce()
                
            
            if right_candidates:
                # Here I need to check for consistency
                choice = softmax_choice(right_candidates,tau = self.tau)
                self.learner.wm.ts2 = TChunk(choice)
                self.add_to_type_dict(pair.s2,choice)
            if not isinstance(self.learner.wm.ts1.structure,list):
                if not self.learner.wm.ts1.structure.is_start():
                    # Here I need to check for consistency
                    if left_candidates:
                        choice = softmax_choice(left_candidates,tau = self.tau)
                        self.learner.wm.ts1 = TChunk(choice)
                        self.add_to_type_dict(pair.s1,choice)
                        
        # This part should try to find a valid starting type
        if not isinstance(self.learner.wm.ts1.structure, list):
            if not self.learner.wm.ts1.structure.is_start():
                if left_candidates:
                    choice = softmax_choice(left_candidates,tau = self.tau)
                    self.learner.wm.ts1 = TChunk(choice)
                    self.add_to_type_dict(pair.s1,choice)
                else:
                    self.learner.wm.ts1 = TChunk(Type.EMPTY)
                    self.add_to_type_dict(pair.s1,Type.EMPTY)
        self.fill_empty_types(pair)
        
        self.correct_typings2(pair)
        
        # Check if s1 is typed (no EMPTY types in ts1)
        # if it is
        # - get candidates type for s2 and bad types for s2
        # - Try to find a ts2 compatible with ts1
        # - if found assign it to ts2
        # - if not found, check if ts1 is expecting something after
        # - if so fulfill expectation if proposed type is not bad
        # - otherwise type ts2 as empty (failure to type)
        # if s1 is not typed:
            # collect candidates types for both s1 and s2.
            # Try to find find compatible types and assign the corresponding types to ts1 and ts2
            # Special cases, only s1 has good types or only s2 have good types. 
            # In that case, chose randomly a type for s1 or s2, if it expects something in the other position, fullfil expectation otherwise failure to type
        #self.learner.wm.ts2 = TChunk(Type.EMPTY)
                
    def correct_typings2(self, pair: ChunkPair):
        if not self.learner.wm.ts1.has_empty_elements() and not self.learner.wm.ts2.has_empty_elements():
            if not isinstance(self.learner.wm.ts1.structure,list) and not isinstance(self.learner.wm.ts2.structure, list):
               # print('Both non complex')
               t1 = self.learner.wm.ts1.structure
               t2 = self.learner.wm.ts2.structure
               if t1.is_expecting_after() and not t2.is_expecting_before():
                   # print('t1 expectations')
                   # Check compatibility and correct if needed
                   # print(f't1 {t1} is expecting after {Type(t1.right_type())} and t2 is {t2}')
                   t1_r = Type(t1.right_type())
                   if t1_r == t2:
                       pass
                       # print('Good match')
                   else:                   
                       # print('Bad match')
                       # Check if expectation is a bad match for t2
                       if t1 in self.learner.ltm.chunk_type_associations[pair.s1]:
                           value_ts1 = self.learner.ltm.chunk_type_associations[pair.s1][t1]
                       else:
                           value_ts1 = 0
                       #print("value 2",value_ts2)                      
                       if t2 in self.learner.ltm.chunk_type_associations[pair.s2]:
                           value_ts2 = self.learner.ltm.chunk_type_associations[pair.s2][t2]
                       else:
                           value_ts2 = 0
                       #print("value 2",value_ts2)
                       candidates = {'s1': value_ts1,'s2': value_ts2}
                       dominant_side = softmax_choice(candidates, tau=self.tau)
                       #print("dominant side",dominant_side)
                       
                       bad_t2 = self.extract_bad_types(pair.s2)
                       good_t2 = self.extract_good_types(pair.s2)
                       if t1_r in bad_t2 or dominant_side == "s2": # or t1_r has not been used for that element
                           # print('Should retype expectation')
                           new_t1 = t1 + t1_r
                           [new_t1,t2] = new_t1.split(pu=0,prim=t2)
                           self.replace_in_type_dict(pair.s1, self.learner.wm.ts1, new_t1)
                           self.replace_in_type_dict(pair.s2, self.learner.wm.ts2, t2)
                           self.learner.wm.ts1 = TChunk(new_t1)
                           self.learner.wm.ts2 = TChunk(t2)
                       else:
                           self.replace_in_type_dict(pair.s2, self.learner.wm.ts2, t1_r)
                           self.learner.wm.ts2 = TChunk(t1_r)
                       # retype here
               elif not t1.is_expecting_after() and t2.is_expecting_before():
                   # print('t2 expectations')
                   # print(f't2 {t2} is expecting before {Type(t2.left_type())} and t1 is {t1}')
                   # Check compatibility and correct if needed
                   t2_l = Type(t2.left_type())
                   if t2_l == t1:
                       pass
                       #print('Good match')
                   else:
                       # print('Bad match')
                       # Check if expectation is a bad type for t1
                       if t1 in self.learner.ltm.chunk_type_associations[pair.s1]:
                           value_ts1 = self.learner.ltm.chunk_type_associations[pair.s1][t1]
                       else:
                           value_ts1 = 0
                       #print("value 2",value_ts2)                      
                       if t2 in self.learner.ltm.chunk_type_associations[pair.s2]:
                           value_ts2 = self.learner.ltm.chunk_type_associations[pair.s2][t2]
                       else:
                           value_ts2 = 0
                       #print("value 2",value_ts2)
                       candidates = {'s1': value_ts1,'s2': value_ts2}
                       dominant_side = softmax_choice(candidates, tau=self.tau)
                       #print("dominant side",dominant_side)
                       bad_t1 = self.extract_bad_types(pair.s1)
                       good_t1 = self.extract_good_types(pair.s1)
                       if t2_l in bad_t1 or dominant_side == "s1":
                           # print('Should retype expectation')
                           new_t2 = t2_l+t2
                           [t1,new_t2] = new_t2.split(pu=1,prim=t1)
                           self.replace_in_type_dict(pair.s2, self.learner.wm.ts2, new_t2)
                           self.replace_in_type_dict(pair.s1, self.learner.wm.ts1, t1)
                           self.learner.wm.ts2 = TChunk(new_t2)
                           self.learner.wm.ts1 = TChunk(t1)
                       else:
                           self.replace_in_type_dict(pair.s1, self.learner.wm.ts1, t2_l)
                           self.learner.wm.ts1 = TChunk(t2_l)
                           # print(t1)
                           # print(new_t2)
                           # add t2 to is left type and split it using t1
                       # retype here
                   pass
               elif t2.is_expecting_before() and t1.is_expecting_after():
                   # Incompatible types! Try to find a compatible pairing
                   print(f'Incompatible typing at step {self.learner.n_reinf}')
                   # retype here
               
            elif self.learner.wm.ts1.is_consistent():
               #print(self.learner.wm.get_responses())
               # print('ts1 complex')
               reduced_type = self.learner.wm.ts1.reduce()
               t2 = self.learner.wm.ts2.structure
               if reduced_type.is_expecting_after() and not t2.is_expecting_before():
                   rt_r = Type(reduced_type.right_type())
                   if rt_r == t2:
                       pass
                       # print('Good match')
                   else:                   
                       # print('Bad match')
                       # Check if expectation is a bad match for t2
                       #unusable_index, type_to_modify = self.learner.wm.ts1.find_type_to_modify()
                       #print("type to modify",type_to_modify)
                       #print("t1",self.learner.wm.ts1.structure,"det indexet jobbar på!",self.learner.wm.ts1.remove_structure2())#type to modify",skiti)                    
                       right_types = self.learner.wm.ts1.right_types()
                       #print("right types",right_types)
                       for i,t in enumerate(right_types):
                           if t.is_expecting_after():
                               valueindex = i
                       #print("value index",valueindex)        
                       right_values = self.learner.wm.get_right_values(pair)
                       #print("right values" ,right_values)
                       competing_t1_value = right_values[valueindex] 
                       #print("competing value", competing_t1_value)
                       #print("pair",pair, "s2",pair.s2)
                       #print("reduced type",reduced_type)
                       #print("chunk-type-associations for s2",self.learner.ltm.chunk_type_associations[pair.s2])
                       #print("t2",t2)
                       if t2 in self.learner.ltm.chunk_type_associations[pair.s2]:
                           value_ts2 = self.learner.ltm.chunk_type_associations[pair.s2][t2]
                       else:
                           value_ts2 = 0
                       #print("value 2",value_ts2)
                       candidates = {'s1': competing_t1_value,'s2': value_ts2}
                       dominant_side = softmax_choice(candidates, tau=self.tau)
                       #print("dominant side",dominant_side)
                       bad_t2 = self.extract_bad_types(pair.s2)
                       good_t2 = self.extract_good_types(pair.s2)
                       if rt_r in bad_t2 or dominant_side == 's2': # or t1_r has not been used for that element
                           # print('Should retype expectation')
                           
                           # print(f'ts1 is {self.learner.wm.ts1} and ts2 is {self.learner.wm.ts2}')
                           if t2.is_primitive():# and len(pair.s1) <=2:
                               # print(f't2 {t2} is a primitive type')
                               oldtyp = self.learner.wm.ts1
                               self.learner.wm.ts1 = self.learner.wm.ts1.retype_expectation(t2,self.learner.wm.get_responses())
                               self.replace_in_type_dict(pair.s1, oldtyp, self.learner.wm.ts1)
                               # print(f'ts1 consistent after changing expectation: {self.learner.wm.ts1.is_consistent()}')
    
                           # print('retyping')
                           # print(f'ts1 is {self.learner.wm.ts1} and ts2 is {self.learner.wm.ts2}')
                           
                           # new_t1 = t1 + t1_r
                           # [new_t1,t2] = new_t1.split(pu=0,prim=t2)
                           # self.learner.wm.ts1 = TChunk(new_t1)
                           # self.learner.wm.ts2 = TChunk(t2)
                       else:
                           self.replace_in_type_dict(pair.s2, self.learner.wm.ts2, rt_r)
                           self.learner.wm.ts2 = TChunk(rt_r)
                       #self.learner.wm.ts2 = TChunk(new_ts2)
                   # ts1 complex: check if it reduces to something that expect something after. If ts2 expects something before retype ts2

        # !!! add types

    def choose_types(self, typ, s1, s2):
        def _compatible_pair(typ, left_candidates, right_candidates):
            # Tries to find a compatible pair
            chosen_pair = None
            for lt in left_candidates:
                for rt in right_candidates:
                    try:
                        if lt + rt == typ:
                            chosen_pair = (lt, rt)
                            break
                    except TypeError:
                        continue
                if chosen_pair:
                    break
            return chosen_pair
        
        def _dominant_type(typ, left_candidates, right_candidates,s1,s2):
            side, chosen = merged_softmax_choice(left_candidates, right_candidates, tau=self.tau)

            if chosen is None:
                # Fall back to a default random split
                left_type, right_type =typ.split(pu=0.5, 
                                                       prim='New',
                                                       bad_s1= self.extract_bad_types(s1),
                                                       bad_s2=self.extract_bad_types(s2))
                return (left_type, right_type)
            elif side == 'left':
                left_type, right_type = typ.split(pu=1.0, prim=chosen,bad_s2 = self.extract_bad_types(s2))
                return (left_type, right_type)
            else:
                left_type, right_type = typ.split(pu=0.0, prim=chosen,bad_s1 = self.extract_bad_types(s1))
                return (left_type, right_type)

        
        left_candidates = self.extract_good_types(s1)
        right_candidates = self.extract_good_types(s2)
        
        chosen_pair = _compatible_pair(typ,left_candidates,right_candidates)
        
            
        if not chosen_pair: # No compatible pairs have been found
                # Choose randomly a right of a left type that is good and construct the corresponding type on the other side
            chosen_pair = _dominant_type(typ,left_candidates,right_candidates,s1,s2)
            
        return chosen_pair
    

    def modify_expectation_at_level(ts, winner_index, new_type, side="right"):
    # """
    # Modify expectation at a given node inside ts (compound or simple).
    # Ensures global consistency by re-splitting from the top.
    
    # Args:
    #     ts: TChunk (possibly compound)
    #     winner_index: index in flattened structure of node to update
    #     new_type: Type to assign as expectation
    #     side: "right" (is_expecting_after) or "left" (is_expecting_before)
    # """
        flat_nodes = ts.flatten_structure()
        target_node = flat_nodes[winner_index]
    
        if side == "right":
            updated = target_node + new_type
            [updated, _] = updated.split(pu=0, prim=new_type)
    
        elif side == "left":
            updated = new_type + target_node
            [_, updated] = updated.split(pu=1, prim=new_type)
    
        else:
            raise ValueError("side must be 'right' or 'left'")
    
        # Replace modified node
        flat_nodes[winner_index] = updated
    
        # Rebuild structure fully to ensure consistency
        rebuilt = TChunk.from_flat_nodes(flat_nodes)
        rebuilt = rebuilt.reduce()  # optional: re-reduce top level to enforce coherence
    
        return rebuilt
    
    
    def propagate_types(self,
                        chunk: SChunk,
                        current_type: Type) -> None:
        
        self.learner.wm.typing_events[chunk] = current_type
    
        if not isinstance(chunk.structure, list):
            return  # Base case: it's a leaf
        
        # Propagate to children
        left_chunk = chunk.get_left()
        right_chunk = chunk.get_right()
        
        
        left_type, right_type = self.choose_types(current_type,left_chunk,right_chunk) #choose_types or choose_types_greedy is for old or new way
        
        self.propagate_types(left_chunk, left_type)
        self.propagate_types(right_chunk, right_type)

    
    def type_sentence(self, s1: SChunk):
        typ = Type.SENTENCE # start with the sentence type
        # Get good and bad types for the components of s1 if its structure is complex
        self.learner.wm.typing_events = {}
        self.propagate_types(s1,typ)
        
    def is_new(self,typ:Type, chunk: SChunk):
        if chunk in self.learner.ltm.chunk_type_associations:
            if typ in self.learner.ltm.chunk_type_associations[chunk]:
                return True
            else:
                return False
        else:
            return False
        
    
    def extract_bad_types(self, chunk: SChunk):
        def filter_dict_below_threshold(data,threshold):
            result = {k: v for k, v in data.items() if v < threshold}
            return result if result else {}
        
        if chunk in self.learner.ltm.chunk_type_associations:
            return filter_dict_below_threshold(self.learner.ltm.chunk_type_associations[chunk],self.bad_type_threshold)
        else:
            return {}
        
    def extract_good_types(self, chunk: SChunk):
        def filter_dict_above_threshold(data,threshold):
            result = {k: v for k, v in data.items() if v > threshold}
            return result if result else {}
        
        if chunk in self.learner.ltm.chunk_type_associations:
            return filter_dict_above_threshold(self.learner.ltm.chunk_type_associations[chunk],self.good_type_threshold)
        else:
            return {}
        
    def extract_good_starting_types(self, chunk: SChunk):
        def filter_dict_above_threshold(data,threshold):
            result = {k: v for k, v in data.items() if v > threshold and k.is_start()}
            return result if result else {}
        
        if chunk in self.learner.ltm.chunk_type_associations:
            return filter_dict_above_threshold(self.learner.ltm.chunk_type_associations[chunk],self.good_type_threshold)
        else:
            return {}
        
    def fill_empty_types(self, pair):
        if not isinstance(self.learner.wm.ts1.structure, list): # ts1 is not complex
            if self.learner.wm.ts1.has_empty_elements() and not self.learner.wm.ts2.has_empty_elements() and self.learner.wm.ts2.structure.is_expecting_before():                
                new_ts1 = Type(self.learner.wm.ts2.structure.left_type())

                self.add_to_type_dict(pair.s1, new_ts1)
                self.learner.wm.ts1 = TChunk(new_ts1)

            elif not self.learner.wm.ts1.has_empty_elements() and self.learner.wm.ts1.structure.is_expecting_after() and self.learner.wm.ts2.has_empty_elements():
                new_ts2 = Type(self.learner.wm.ts1.structure.right_type())
                self.add_to_type_dict(pair.s2, new_ts2)
                self.learner.wm.ts2 = TChunk(new_ts2)
        elif self.learner.wm.ts1.is_consistent():
            reduced_type = self.learner.wm.ts1.reduce()
            if reduced_type.is_expecting_after() and self.learner.wm.ts2.has_empty_elements():
                print('This case applies')
                new_ts2 = Type(reduced_type.right_type())
                self.add_to_type_dict(pair.s2, new_ts2)
                self.learner.wm.ts2 = TChunk(new_ts2)
            # Here I need to check whether the reduced ts1 is expecting something and at which level.
            # Will be implemented later. I may use the full reduction as a first step.
            pass

        
        # if ts1 not a list and empty and ts2 non empty and expecting before
        # assign expectation to ts1
        # elif ts1 is consistent and expecting after and ts2 empty
        # assign expectation to ts2




import tracemalloc
class Learner():
    ID = 0
    
    def __init__(self, config: LearnerConfig,samplenum,cfg,output=False,clustering=False):
        self.ID = Learner.ID + 1
        Learner.ID +=1
        
        self.ltm = LongTermMemory(config)
        # print(type(self), type(config), type(samplenum), type(cfg))
        self.wm = WorkingMemory(self,config,samplenum,cfg,output=output,clustering=clustering)
        self.history = LearningHistory()
        self.chaining = config.chaining
        self.type_on = config.type_on
        
        
        # self.border_type = border # or 'default'
        self.n_trials = config.n_trials
        self.n_reinf = 0
        
        self.final_index = 0
        
        
        
    def __repr__(self):
        return f'Learner {self.ID}'
       
    def learn(self,stimuli_stream):
        # initialize stimuli
        w = Word(stimuli_stream.read_stimuli(0), self.wm.db)
        s1 = SChunk(w.t)
        self.wm.ws = [w]

        s2_index = 1
        milestones = [int(self.n_trials*x/100) for x in range(1,102)]
        i = 0
        while self.n_reinf <= self.n_trials:
            if self.n_reinf == milestones[i]:
                print(" "*(int(100*self.n_reinf/self.n_trials)-1) + f"{int(100*self.n_reinf/self.n_trials)}%")
                i+=1

                # snapshot = tracemalloc.take_snapshot()
                # top_stats = snapshot.statistics('lineno')
                # print("Top 10 memory-consuming lines:")
                # for stat in top_stats[:10]:
                #     print(stat)

            if not self.chaining:
                if self.type_on:
                    s1, s2_index = self.wm.respond_with_type(stimuli_stream, s1, s2_index)
                else:
                    s1, s2_index = self.wm.respond(stimuli_stream, s1, s2_index)
            else:
                s1, s2_index = self.wm.respond_with_chaining2(stimuli_stream, s1, s2_index)

        self.final_index = s2_index
        
        if self.wm.do_output(): self.wm.db.output()

class Dir:
    L = 1
    R = 2